import io
import openpyxl
from PIL import Image


def create_book(name_book='Result.xlsx'):
    book = openpyxl.Workbook()
    book.save(name_book)


def add_info_to_book(name_sheets, name_photo, num_row, result):
    wb = openpyxl.load_workbook('Result.xlsx')
    sheet_list = wb.sheetnames
    if 'Sheet' in sheet_list:
        wb.remove(wb.active)
    if name_sheets not in sheet_list:
        wb.create_sheet(name_sheets)

    ws = wb[name_sheets]

    ws.column_dimensions['A'].width = 22
    ws.row_dimensions[num_row].height = 153

    img = Image.open(name_photo).resize((200, 200))
    buf = io.BytesIO()
    img.save(buf, format='png')
    img = openpyxl.drawing.image.Image(buf)

    img.anchor = f'A{num_row}'

    ws.add_image(img)
    ws[f"B{num_row}"] = name_photo
    ws.column_dimensions['B'].width = len(name_photo)
    ws[f"C{num_row}"] = result
    wb.save('Result.xlsx')


if __name__ == "__main__":
    pass
    # create_book()
    # add_info_to_book('blogger', 'blogger.jpg', 1, 'True')
    # add_info_to_book('placeman', 'placename.jpg', 2, "False")
